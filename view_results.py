
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import plotly.express as px
from pathlib import Path

st.set_page_config(page_title="Prompt Matrix Viewer", layout="wide")

st.title("📊 Prompt Matrix Viewer with Searchable Company Dropdown")

EXCEL_FILE = "matrix_template2.xlsm"

if not Path(EXCEL_FILE).exists():
    st.error(f"'{EXCEL_FILE}' not found in current directory.")
    st.stop()

wb = load_workbook(EXCEL_FILE, data_only=True)
sheet = wb.active

# Extract timestamp from cell A1
timestamp = sheet["A1"].value
if timestamp:
    st.caption(f"Last updated: {timestamp}")

# Extract companies from row 16
companies = []
col = 2
while sheet.cell(row=16, column=col).value:
    companies.append(sheet.cell(row=16, column=col).value)
    col += 1

# Extract prompts from column A starting row 18
prompts = []
row = 18
while sheet.cell(row=row, column=1).value:
    prompts.append(sheet.cell(row=row, column=1).value)
    row += 1

# Build DataFrame from matrix
data = []
for i, prompt in enumerate(prompts, start=18):
    row_data = {"Prompt": prompt}
    for j, company in enumerate(companies, start=2):
        row_data[company] = sheet.cell(row=i, column=j).value
    data.append(row_data)

df = pd.DataFrame(data)

# --- Multiselect Dropdown with Search ---
selected_companies = st.multiselect(
    "🔍 Select companies to display:",
    options=companies,
    default=companies,
    help="Type to search and select one or more companies to filter the results table."
)

# Filtered dataframe
filtered_df = df[["Prompt"] + selected_companies]
st.dataframe(filtered_df, use_container_width=True)

# --- Text Length Heatmap ---
with st.expander("📏 Text Length Heatmap (Selected Companies Only)"):
    length_df = filtered_df.copy()
    for col in selected_companies:
        length_df[col] = length_df[col].apply(lambda x: len(str(x)) if x else 0)

    fig = px.imshow(length_df[selected_companies].astype(int),
                    labels=dict(x="Company", y="Prompt #", color="Char Count"),
                    x=selected_companies,
                    y=[f"#{i+1}" for i in range(len(prompts))],
                    text_auto=True,
                    aspect="auto",
                    color_continuous_scale="Viridis")
    st.plotly_chart(fig, use_container_width=True)