import json
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from datetime import datetime

# === Load Excel Workbook ===
filename = "matrix_template2.xlsm"
wb = load_workbook(filename, data_only=True, keep_vba=True)
sheet = wb.active

# === Configuration Setup ===
# Maps config keys to their corresponding row in the sheet (column B is default/global)
config_row_map = {
    "API-key": 4,
    "Model": 5,
    "Temperature": 6,
    "Max-Token": 7,
    "Top_P": 8,
    "Frequency_Penalty": 9,
    "Presence_Penalty": 10,
    "Systemprompt": 12
}

# === Normalize function for comparison ===
def normalize(val):
    if val is None:
        return ""
    return str(val).strip().replace(",", ".").lower()

# === Load Global Config from Column B ===
global_config = {
    key: sheet.cell(row=row, column=2).value for key, row in config_row_map.items()
}

# === Initialize OpenAI Client ===
client = OpenAI(api_key=global_config["API-key"])

# === Read Company Columns from Row 16 ===
companies = []
col = 2
while sheet.cell(row=16, column=col).value:
    companies.append((col, sheet.cell(row=16, column=col).value))
    col += 1

# === Read Prompts from Column A (Row 18 onwards) ===
prompts = []
row = 18
while sheet.cell(row=row, column=1).value:
    prompts.append((row, sheet.cell(row=row, column=1).value))
    row += 1

# === Open log file ===
with open("generation_log.txt", "w", encoding="utf-8") as log_file:
    log_file.write(f"Prompt generation log - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

    # === Process Each Prompt for Each Company ===
    for r, prompt in prompts:
        for c, company in companies:
            # Skip if cell already has a value
            if sheet.cell(row=r, column=c).value not in (None, ""):
                continue

            # Load column-specific config or fallback to global config
            col_config = {
                key: sheet.cell(row=row_num, column=c).value or global_config[key]
                for key, row_num in config_row_map.items()
            }

            # Highlight overrides where column differs from default (normalized comparison)
            for key, row_num in config_row_map.items():
                user_val = normalize(sheet.cell(row=row_num, column=c).value)
                default_val = normalize(sheet.cell(row=row_num, column=2).value)
                if user_val != default_val:
                    sheet.cell(row=row_num, column=c).fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

            # Parse model parameters
            max_tokens = int(col_config["Max-Token"])
            model = col_config["Model"]
            temperature = float(str(col_config["Temperature"]).replace(",", "."))
            top_p = float(str(col_config["Top_P"]).replace(",", "."))
            frequency_penalty = float(col_config["Frequency_Penalty"])
            presence_penalty = float(col_config["Presence_Penalty"])
            context = col_config.get("Systemprompt", "")
            max_tokens = max_tokens

            # Combine prompt
            full_prompt = f"{context}\n\nCompany: {company}\n\nTask: {prompt}"

            # Call OpenAI API
            response = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": full_prompt}],
                temperature=temperature,
                top_p=top_p,
                frequency_penalty=frequency_penalty,
                presence_penalty=presence_penalty
            )

            result = response.choices[0].message.content.strip()
            if result.startswith("=") or result.startswith("-"):
                result = "'" + result  # Prevent Excel formula interpretation

            # Write result to cell
            cell = sheet.cell(row=r, column=c)
            cell.value = result
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            # Optional keyword-based cell highlighting
            if "risk" in result.lower():
                cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            elif "compliance" in result.lower():
                cell.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")

            # Write to log (plain text only)
            log_file.write(f"R{r}C{c} ({company}) - Prompt: '{prompt[:40]}...' completed.\n")

# === Timestamp update ===
sheet["A1"] = f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

# === Auto-adjust column widths ===
for c, _ in companies:
    max_length = 0
    for r, _ in prompts:
        val = sheet.cell(row=r, column=c).value
        if val:
            max_length = max(max_length, len(str(val)))
    col_letter = chr(64 + c) if c <= 26 else "A" + chr(64 + c - 26)
    sheet.column_dimensions[col_letter].width = min(max_length * 0.9, 60)

# === Save workbook ===
try:
    wb.save(filename)
    print(f"File updated: {filename}")
except PermissionError:
    fallback = "matrix_template2_updated.xlsm"
    wb.save(fallback)
    print(f"File locked. Saved as: {fallback}")
