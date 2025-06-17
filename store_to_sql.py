import json
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine

# Load config
with open("config.json") as f:
    config = json.load(f)

db = config["database"]
table = config["table"]

# Load Excel
wb = load_workbook("matrix_template2.xlsm", data_only=True)
sheet = wb.active

# Collect companies from row 16
companies = []
col = 2
while sheet.cell(row=16, column=col).value:
    companies.append((col, sheet.cell(row=16, column=col).value))
    col += 1

# Collect prompts and results
rows = []
row = 17
while sheet.cell(row=row, column=1).value:
    prompt = sheet.cell(row=row, column=1).value
    for c, company in companies:
        result = sheet.cell(row=row, column=c).value
        if result:
            rows.append({"prompt": prompt, "company": company, "response": result})
    row += 1

df = pd.DataFrame(rows)
engine = create_engine(f"sqlite:///{db}")
df.to_sql(table, con=engine, if_exists="append", index=False)
print(f"✅ {len(df)} rows stored to {table} in {db}")
